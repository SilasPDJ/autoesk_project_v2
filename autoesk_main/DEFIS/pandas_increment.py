import pandas
import openpyxl
import os

from autoesk_main._new_set_paths import NewSetPaths
from autoesk_main.default.data_treatment import ExcelToData


class Pandas(NewSetPaths, ExcelToData):

    my_file = 'DEFIS-anual-Copia.xlsx'

    def __init__(self, compt_file=None):
        # sh_names = ['DEFIS']

        excel_path = os.path.dirname(super().excel_file_path())
        self.my_file = f'{excel_path}/{self.my_file}'

    def os_walk__get_dirfs_path(self, searched_client=None):
        """
        :param searched_client: None or False => return all paths,
        :return:
        """
        clientes_path = {}

        INITIAL_PATH = r'I:\OESK_CONTABIL'
        ano_dirf = str(self.y())

        for (dirpath, dirnames, filenames) in os.walk(INITIAL_PATH):
            if ano_dirf in dirnames:
                # get last part of folder name ↓↓↓
                the_client = os.path.basename(os.path.normpath(dirpath))
                    # nome tem que ser exato

                for (dp2, dn2, fn2) in os.walk(dirpath):
                    if ano_dirf in dp2:
                        for dp3, dn3, fn3 in os.walk(dp2):
                            if 'DIRF' in dp3:
                                dirf_path = dp3
                                now_command = f'explorer {dp3}'
                                clientes_path[the_client] = dirf_path
                                # call(now_command)
        for klient, vdir in clientes_path.items():
            if klient == searched_client:
                return klient
            yield vdir


inst = Pandas()
input('security')
mcp = openpyxl.load_workbook(inst.my_file)
# mycopy.create_sheet('Socios')

wks = mcp.worksheets

ws1 = wks[0]
ws2 = wks[1]
ws3 = wks[2]

# input(f1st['A1'].value)
"""
for e in sheet1['A1:E20']:
    for i in range(10):
        print(f'A{i}')
"""
__dict = {}
campos = []


# Esse loop ele faz de acordo com o que trabalho no Simples Nacional, vou ler o dicionário e
# incrementar na planilha de sócios as informações tais como CNPJ, principalmente
for iter_row in ws3.iter_rows(max_col=8):
    for cell in iter_row:
        row, coluna = cell.row, cell.column
        valor = cell.value
        valor = str(valor).strip()
        if row == 1:
            __dict[valor] = []
            campos.append(valor)

        else:
            __dict[campos[coluna-1]].append(valor)

print('\033[1;33mDICT...\033[m')
for k, list_values in __dict.items():

    print(k, ': ')
    for val in list_values:
        print(val)
    # input(len(list_values))


for val in inst.os_walk__get_dirfs_path():
    print(val)



input(ws2['A1'].value)

mcp.save(inst.my_file)
