import os
from time import sleep
from default.settings import SetPaths
from default.data_treatment import ExcelToData

from default.data_treatment import transformers as tfms
import openpyxl



# pdf2jpg()
# jpg2txt()


class VisualizaTicket(SetPaths, ExcelToData):

    contador = 1

    def __init__(self, compt_file=None):
        """
        :param compt_file: from GUI

        # remember past_only arg from self.get_atual_competencia
        """
        import pandas as pd
        from default.webdriver_utilities.pre_drivers import pgdas_driver

        # O vencimento DAS(seja pra qual for a compt) está certo, haja vista que se trata do mes atual

        sh_names = ['DEFIS']
        if compt_file is None:
            compt_file = self.compt_and_filename()
            compt, excel_file_name = compt_file
        else:
            compt, excel_file_name = compt_file

        COMPT = compt = f"DEFIS_{self.y()}"
        # transcrevendo compt para que não seja 02/2021

        # excel_file_name = '/'.join(excel_file_name.split('/')[:-1])
        excel_file_name = os.path.dirname(excel_file_name)
        excel_file_name += f'/DEFIS-anual.xlsx'
        pdExcelFile = pd.ExcelFile(excel_file_name)

        for sh_name in sh_names:
            # agora eu posso fazer downloalds sem me preocupar tendo a variável path

            msh = pdExcelFile.parse(sheet_name=str(sh_name))
            col_str_dic = {column: str for column in list(msh)}

            msh = pdExcelFile.parse(sheet_name=str(sh_name), dtype=col_str_dic)
            READ = self.le_excel_each_one(msh)
            self.after_READ = self.readnew_lista(READ, False)
            after_READ = self.after_READ

            for i, CNPJ in enumerate(after_READ['CNPJ']):
                # ####################### A INTELIGENCIA EXCEL ESTÁ SEM OS SEM MOVIMENTOS NO MOMENTO
                _cliente = after_READ['Razão Social'][i]
                _ja_declared = after_READ['Declarado'][i].upper().strip()
                _cod_sim = after_READ['Código Simples'][i]
                _cpf = after_READ['CPF'][i]
                _cert_or_login = after_READ['CERTORLOGIN'][i]

                # Dirfis exclusivos search
                _dirf_sch = after_READ['DIRF'][i]

                # self.dirf_nome = _dirf_sch if _dirf_sch != '-' else _cliente
                self.dirf_nome = CNPJ

                if _cliente == '':
                    break

                if _ja_declared not in ['S', 'OK', 'FORA']:
                    # ############################################################################################ ↓
                    # self.client_path = self._files_path_defis(_cliente, tup_path=(COMPT, excel_file_name))
                    self.client_path = self._files_path_defis(_cliente, tup_path=(COMPT, excel_file_name))

                    file_pdf = 'VisualizaTicket.pdf'

                    dir_searched_now = self.client_path
                    file_src = ''.join([os.path.join(dir_searched_now, fname) for fname in os.listdir(dir_searched_now) if
                                        fname == file_pdf])
                    #  file_searched = os.path.basename(os.path.normpath(file_searched_path))
                    if file_src != '':
                        now_txts = tfms.pdf2txt(file_src)
                        # Consegui finalmente essa maravilha...
                        self.here_scrap(now_txts)

    def here_scrap(self, txt):

        pasinit = txt

        user = splitxt = txt.split()
        USED = ' '.join(user)

        # input(used)

        backup1 = USED

        # ############################################## caças responsivas
        # ####### caça CPFs

        cpfs = []
        for e, cpf in enumerate(USED.split()):
            cpf = self.str_with_mask(cpf, '000.000.000-00')
            if cpf:
                cpfs.append(cpf)

        cpfs = list(dict.fromkeys(cpfs))
        cotas = []
        # preciso fazer o set pois ele vai usar os ultimos
        for v1, v2 in zip(range(0, len(USED)), range(20, len(USED))):
            if v2 == len(USED):
                break
            if '$'.lower() in USED[v1].lower():
                val = USED[v1:v2]
                try:
                    val = val[:val.index(',')+3]

                    # cotas.add(val)
                    cotas.append(val)
                except ValueError:
                    pass

        if 'RETIRA-SE' in splitxt:
            retirado = cpfs[-1]
            cpfs = cpfs[:-1]

        dalecotas = [f.split()[1] for f in cotas]
        dalecotas = [f.replace('.', '') for f in dalecotas]
        dalecotas = [f'R$ {r}' for r in dalecotas]
        print(cpfs)
        print(self.dirf_nome)
        print('-'*30)
        if len(cpfs) > 3:
            cpfs = cpfs[-2:]
        for cpf, cota in zip(cpfs, dalecotas):
            self.contador += 1
            ws2.cell(self.contador, 1).value = self.dirf_nome
            ws2.cell(self.contador, 2).value = cpf
            ws2.cell(self.contador, 3).value = cota

        print(dalecotas)

        print('-'*30)


my_file = os.path.dirname(SetPaths().compt_and_filename()[1])
my_file += f'/DEFIS-anual-Copia.xlsx'

mcp = openpyxl.load_workbook(my_file)
wks = mcp.worksheets
ws1 = wks[0]
ws2 = wks[1]
ws3 = wks[2]

VisualizaTicket()


# mcp.save(os.path.dirname(SetPaths().compt_and_filename()[1])+'\\teste.xlsx')
mcp.save(my_file)
