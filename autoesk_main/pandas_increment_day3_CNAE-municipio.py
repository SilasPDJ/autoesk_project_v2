import pandas
import openpyxl
import os

from autoesk_main._new_set_paths import NewSetPaths
from autoesk_main.imports import ExcelToData
from autoesk_main.imports import WDShorcuts
from autoesk_main.imports import ginfess_driver
from bs4 import BeautifulSoup
import requests


# TODO
# EXECUTAR ESSE SCRIPT


class Pandas(NewSetPaths, ExcelToData):

    my_file = 'DEFIS-anual-Copia.xlsx'

    def __init__(self):
        # sh_names = ['DEFIS']

        excel_path = os.path.dirname(super().excel_file_path())
        self.my_file = f'{excel_path}\\{self.my_file}'

inst = Pandas()

mcp = openpyxl.load_workbook(inst.my_file)
# mycopy.create_sheet('Socios')

wks = mcp.worksheets

ws1 = wks[0]
ws2 = wks[1]

# input(f1st['A1'].value)
"""
for e in sheet1['A1:E20']:
    for i in range(10):
        print(f'A{i}')
"""
__dict = {}
campos = []


# Esse loop ele faz de acordo com o que trabalho no Simples Nacional, vou ler o dicionário e
# incrementar na planilha de sócios as informações tais como CNPJ, principalmente

for iter_row in ws1.iter_rows(min_col=1, max_col=2):
    for cell in iter_row:
        row, coluna = cell.row, cell.column
        valor = cell.value
        valor = str(valor).strip()
        if row == 1:
            __dict[valor] = []
            campos.append(valor)

        else:
            __dict[campos[coluna-1]].append(valor)

"""print('\033[1;33mDICT...\033[m')
for k, list_values in __dict.items():

    print(k, ': ')
    for val in list_values:
        print(val)
    # input(len(list_values))
"""


print(*__dict["CNPJ"])
cnpjs = __dict["CNPJ"]


# detection ##############################################################################################


# ########################################################################################################

def soup(me):
    """
    :param me: element
    :return:
    """
    me = str(me)
    btf = BeautifulSoup(me, 'html.parser')
    return btf


PARTIR_DE = 15
# driver = ginfess_driver()
for cont, cnpj in enumerate(cnpjs[PARTIR_DE:]):
    try:
        import re
        # CADASTRA CNAE / DESCRIÇÃO
        now_link = f'http://cnpj.info/{cnpj}'
        # driver.get(now_link)
        # cnae = driver.find_element_by_tag_name("u").text

        req = requests.get(now_link).text
        thesoup = soup(req)
        """
        # o select gera uma lista, aceitando, indexação ou for... etc
        cnae_total = thesoup.select("u")[0].text

        # cnae_total significa só "-"
        # RSPLIT
        cnae, descricao = cnae_total.rsplit(cnae_total[11], 1)
        cnae = cnae.strip()
        descricao = descricao.strip()
        print(cnae, descricao, '------>', cnpj)

        ws1[f"J{str(cont+2)}"].value = cnae
        ws1[f"K{str(cont+2)}"].value = descricao
        """
        # endereço
        # document.querySelector("body > div.container > div > div > div.col.c9-2 > p:nth-child(14)")
        endereco = thesoup.find('h3', text=re.compile('Endereço'),) #attrs={'class': 'pos'})
        try:
            items = endereco.next_siblings
        except AttributeError:
            print('teste atr')
            continue
        print('teste')
        _dados_header = ['Logadouro', 'Complemento', 'Bairro', 'Cidade e Estado', 'CEP']
        dados = []
        for e, txt in enumerate(items):
            txt = str(txt).strip()
            if e == 9:
                break
            if 'br' not in txt:
                print(txt)
                dados.append(txt) if txt != '' else None

        if '<h3>Contatos</h3>' == dados[-1]:
            dados.insert(1, '-')

        dados = {k: v for k, v in zip(_dados_header, dados)}
        try:
            dados["Cidade"], dados["Estado"] = dados['Cidade e Estado'].split('-')
        except ValueError:
            input(f"NÃO DEU CERTO, LINHA dados['Cidade e Estado']: {dados['Cidade e Estado']}")

        del dados["Cidade e Estado"]
        del _dados_header[-2]
        _dados_header.append("Cidade")
        _dados_header.append("Estado")

        for row in ws1.iter_rows(cont+2, cont+2, 12, 17):
            cont_cell = 0
            for cell in row:
                # if cont_cell == 3:
                try:
                    cell.value = dados[_dados_header[cont_cell]]
                except KeyError:
                    input(f'inputing: {dados}')
                cont_cell += 1


            # el = thesoup.find(f"body > div.container > div > div > div.col.c9-2 > p:nth-child({i})")
        # print(el[0].text)
    except IndexError:
        raise IndexError


mcp.save(inst.my_file)
print('finish')
